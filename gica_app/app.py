from flask import (Flask, render_template, request, redirect, url_for,
                   flash, jsonify, session, g, send_file, make_response)
import sqlite3
import os
import io
import csv as csv_module
from datetime import datetime
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '.env'))
except ImportError:
    pass

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

app = Flask(__name__)
app.secret_key = os.environ.get('GICA_SECRET_KEY', 'gica$secretaria#salud@2026_XK9!')
DB_PATH = os.environ.get('DB_PATH', os.path.join(os.path.dirname(__file__), 'gica.db'))


def _ensure_db():
    """Inicializa la BD automáticamente si no existe y aplica migraciones."""
    import subprocess, sys
    if not os.path.exists(DB_PATH):
        init_script = os.path.join(os.path.dirname(__file__), 'init_db.py')
        subprocess.run([sys.executable, init_script], check=True)
    migr_script = os.path.join(os.path.dirname(__file__), 'agregar_vigencias.py')
    if os.path.exists(migr_script):
        subprocess.run([sys.executable, migr_script], check=True)


_ensure_db()


# ─── DB ───────────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def get_vigencia_id(db):
    """Retorna vigencia_id de sesión o el de la vigencia activa."""
    vid = session.get('vigencia_id')
    if vid:
        if db.execute('SELECT id FROM vigencia WHERE id=?', (vid,)).fetchone():
            return vid
    vig = db.execute(
        'SELECT id FROM vigencia WHERE activa=1 ORDER BY año DESC LIMIT 1'
    ).fetchone()
    if vig:
        session['vigencia_id'] = vig['id']
        return vig['id']
    return None


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def semaforo(valor):
    if valor is None:
        return 'rojo'
    if valor >= 0.9:
        return 'verde'
    if valor >= 0.7:
        return 'amarillo'
    return 'rojo'


def semaforo_color(valor):
    s = semaforo(valor)
    return {'verde': '#198754', 'amarillo': '#ffc107', 'rojo': '#dc3545'}.get(s, '#6c757d')


app.jinja_env.globals.update(semaforo=semaforo, semaforo_color=semaforo_color)


@app.context_processor
def inject_globals():
    resultado = {
        'now': datetime.now(),
        'session_user': session.get('user'),
        'vigencia_actual': None,
        'todas_vigencias': [],
    }
    if 'user' in session:
        try:
            db = get_db()
            resultado['todas_vigencias'] = db.execute(
                'SELECT * FROM vigencia ORDER BY año DESC'
            ).fetchall()
            vid = session.get('vigencia_id')
            va = None
            if vid:
                va = db.execute('SELECT * FROM vigencia WHERE id=?', (vid,)).fetchone()
            if not va:
                va = db.execute(
                    'SELECT * FROM vigencia WHERE activa=1 ORDER BY año DESC LIMIT 1'
                ).fetchone()
                if va:
                    session['vigencia_id'] = va['id']
            resultado['vigencia_actual'] = va
            db.close()
        except Exception:
            pass
    return resultado


@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404


@app.errorhandler(500)
def server_error(e):
    return render_template('errors/500.html'), 500


def registrar_auditoria(accion, modulo='', detalle=''):
    """Registra una acción en el log de auditoría."""
    if 'user' not in session:
        return
    try:
        db = get_db()
        db.execute('''
            INSERT INTO log_auditoria (usuario_id, username, accion, modulo, detalle, ip)
            VALUES (?,?,?,?,?,?)
        ''', (session['user']['id'], session['user']['username'],
              accion, modulo, detalle,
              request.remote_addr or ''))
        db.commit()
        db.close()
    except Exception:
        pass


# ─── DECORADORES DE ACCESO ───────────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            flash('Debe iniciar sesión para acceder al sistema.', 'warning')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated


def puede_editar(f):
    """Nivel <= 3: Administrador, Líder GICA, Líder de Proceso."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            flash('Debe iniciar sesión.', 'warning')
            return redirect(url_for('login'))
        if session['user']['rol_nivel'] > 3:
            flash('No tiene permisos para realizar esta acción.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def puede_eliminar(f):
    """Nivel <= 2: Administrador, Líder GICA."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            flash('Debe iniciar sesión.', 'warning')
            return redirect(url_for('login'))
        if session['user']['rol_nivel'] > 2:
            flash('No tiene permisos para eliminar registros.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


def solo_admin(f):
    """Nivel == 1: Solo Administrador."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user' not in session:
            flash('Debe iniciar sesión.', 'warning')
            return redirect(url_for('login'))
        if session['user']['rol_nivel'] != 1:
            flash('Esta sección es exclusiva del Administrador.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated


# ─── AUTENTICACIÓN ──────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        db = get_db()
        user = db.execute('''
            SELECT u.*, r.nombre as rol_nombre, r.nivel as rol_nivel,
                   r.puede_crear, r.puede_editar as rol_editar,
                   r.puede_eliminar as rol_eliminar, r.puede_admin_usuarios,
                   p.nombre as proceso_nombre
            FROM usuario u
            LEFT JOIN rol r ON u.rol_id = r.id
            LEFT JOIN proceso p ON u.proceso_id = p.id
            WHERE u.username = ?
        ''', (username,)).fetchone()

        if not user or not check_password_hash(user['password_hash'], password):
            flash('Usuario o contraseña incorrectos.', 'danger')
            db.execute('INSERT INTO log_auditoria (username, accion, modulo, detalle, ip) VALUES (?,?,?,?,?)',
                       (username, 'LOGIN_FALLIDO', 'Auth',
                        f'Intento fallido para usuario: {username}', request.remote_addr or ''))
            db.commit()
            db.close()
            return render_template('auth/login.html')

        if not user['activo']:
            flash('Su cuenta está desactivada. Contacte al administrador.', 'danger')
            db.close()
            return render_template('auth/login.html')

        session['user'] = {
            'id': user['id'],
            'username': user['username'],
            'nombre_completo': user['nombre_completo'],
            'email': user['email'],
            'rol_id': user['rol_id'],
            'rol_nombre': user['rol_nombre'],
            'rol_nivel': user['rol_nivel'],
            'puede_crear': bool(user['puede_crear']),
            'puede_editar': bool(user['rol_editar']),
            'puede_eliminar': bool(user['rol_eliminar']),
            'puede_admin': bool(user['puede_admin_usuarios']),
            'proceso_id': user['proceso_id'],
            'proceso_nombre': user['proceso_nombre'],
        }
        db.execute('UPDATE usuario SET ultimo_acceso=CURRENT_TIMESTAMP WHERE id=?', (user['id'],))
        db.execute('INSERT INTO log_auditoria (usuario_id, username, accion, modulo, ip) VALUES (?,?,?,?,?)',
                   (user['id'], username, 'LOGIN', 'Auth', request.remote_addr or ''))
        vig = db.execute(
            'SELECT id FROM vigencia WHERE activa=1 ORDER BY año DESC LIMIT 1'
        ).fetchone()
        if vig:
            session['vigencia_id'] = vig['id']
        db.commit()
        db.close()

        if user['debe_cambiar_password']:
            flash('Por seguridad, debe cambiar su contraseña antes de continuar.', 'warning')
            return redirect(url_for('cambiar_password'))

        next_url = request.args.get('next')
        return redirect(next_url if next_url and next_url.startswith('/') else url_for('dashboard'))

    return render_template('auth/login.html')


@app.route('/logout')
@login_required
def logout():
    registrar_auditoria('LOGOUT', 'Auth')
    session.clear()
    flash('Sesión cerrada correctamente.', 'info')
    return redirect(url_for('login'))


@app.route('/cambiar-vigencia/<int:vid>', methods=['POST'])
@login_required
def cambiar_vigencia(vid):
    db = get_db()
    v = db.execute('SELECT * FROM vigencia WHERE id=?', (vid,)).fetchone()
    db.close()
    if v:
        session['vigencia_id'] = vid
        flash(f'Vigencia cambiada a: {v["nombre"]}', 'info')
    else:
        flash('Vigencia no encontrada.', 'danger')
    return redirect(request.referrer or url_for('dashboard'))


@app.route('/mi-perfil', methods=['GET', 'POST'])
@login_required
def mi_perfil():
    db = get_db()
    usuario = db.execute('''
        SELECT u.*, r.nombre as rol_nombre, p.nombre as proceso_nombre
        FROM usuario u
        LEFT JOIN rol r ON u.rol_id = r.id
        LEFT JOIN proceso p ON u.proceso_id = p.id
        WHERE u.id=?
    ''', (session['user']['id'],)).fetchone()
    if request.method == 'POST':
        nombre = request.form.get('nombre_completo', '').strip()
        email = request.form.get('email', '').strip()
        db.execute('UPDATE usuario SET nombre_completo=?, email=? WHERE id=?',
                   (nombre, email, session['user']['id']))
        db.commit()
        session['user']['nombre_completo'] = nombre
        session['user']['email'] = email
        registrar_auditoria('EDITAR', 'Perfil', 'Actualizó su perfil')
        flash('Perfil actualizado correctamente.', 'success')
        db.close()
        return redirect(url_for('mi_perfil'))
    logs = db.execute('''
        SELECT * FROM log_auditoria WHERE usuario_id=?
        ORDER BY fecha DESC LIMIT 20
    ''', (session['user']['id'],)).fetchall()
    db.close()
    return render_template('auth/perfil.html', usuario=usuario, logs=logs)


@app.route('/cambiar-password', methods=['GET', 'POST'])
@login_required
def cambiar_password():
    if request.method == 'POST':
        actual = request.form.get('password_actual', '')
        nueva = request.form.get('password_nueva', '')
        confirmar = request.form.get('password_confirmar', '')
        db = get_db()
        user = db.execute('SELECT * FROM usuario WHERE id=?', (session['user']['id'],)).fetchone()
        if not check_password_hash(user['password_hash'], actual):
            flash('La contraseña actual no es correcta.', 'danger')
            db.close()
            return render_template('auth/cambiar_password.html')
        if len(nueva) < 8:
            flash('La nueva contraseña debe tener al menos 8 caracteres.', 'danger')
            db.close()
            return render_template('auth/cambiar_password.html')
        if nueva != confirmar:
            flash('Las contraseñas nuevas no coinciden.', 'danger')
            db.close()
            return render_template('auth/cambiar_password.html')
        db.execute('UPDATE usuario SET password_hash=?, debe_cambiar_password=0 WHERE id=?',
                   (generate_password_hash(nueva), session['user']['id']))
        db.commit()
        registrar_auditoria('CAMBIO_PASSWORD', 'Auth', 'Cambió su contraseña')
        flash('Contraseña cambiada correctamente.', 'success')
        db.close()
        return redirect(url_for('dashboard'))
    return render_template('auth/cambiar_password.html')


# ─── GESTIÓN DE USUARIOS ─────────────────────────────────────────────────────

@app.route('/admin/usuarios')
@solo_admin
def usuarios_list():
    db = get_db()
    usuarios = db.execute('''
        SELECT u.*, r.nombre as rol_nombre, r.nivel as rol_nivel,
               p.nombre as proceso_nombre
        FROM usuario u
        LEFT JOIN rol r ON u.rol_id = r.id
        LEFT JOIN proceso p ON u.proceso_id = p.id
        ORDER BY u.activo DESC, r.nivel, u.nombre_completo
    ''').fetchall()
    db.close()
    return render_template('admin/usuarios_list.html', usuarios=usuarios)


@app.route('/admin/usuarios/nuevo', methods=['GET', 'POST'])
@solo_admin
def usuario_nuevo():
    db = get_db()
    roles = db.execute('SELECT * FROM rol ORDER BY nivel').fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        nombre = request.form.get('nombre_completo', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        rol_id = request.form.get('rol_id')
        proceso_id = request.form.get('proceso_id') or None
        debe_cambiar = 1 if request.form.get('debe_cambiar_password') else 0

        if not username or not nombre or not password or not rol_id:
            flash('Complete todos los campos obligatorios.', 'danger')
            db.close()
            return render_template('admin/usuario_form.html',
                                   usuario=None, roles=roles, procesos=procesos)
        if len(password) < 8:
            flash('La contraseña debe tener al menos 8 caracteres.', 'danger')
            db.close()
            return render_template('admin/usuario_form.html',
                                   usuario=None, roles=roles, procesos=procesos)
        existente = db.execute('SELECT id FROM usuario WHERE username=?', (username,)).fetchone()
        if existente:
            flash(f'El nombre de usuario "{username}" ya existe.', 'danger')
            db.close()
            return render_template('admin/usuario_form.html',
                                   usuario=None, roles=roles, procesos=procesos)
        db.execute('''
            INSERT INTO usuario (username, nombre_completo, email, password_hash,
                                 rol_id, proceso_id, debe_cambiar_password)
            VALUES (?,?,?,?,?,?,?)
        ''', (username, nombre, email, generate_password_hash(password),
              rol_id, proceso_id, debe_cambiar))
        db.commit()
        registrar_auditoria('CREAR', 'Usuarios', f'Creó usuario: {username}')
        flash(f'Usuario "{username}" creado correctamente.', 'success')
        db.close()
        return redirect(url_for('usuarios_list'))
    db.close()
    return render_template('admin/usuario_form.html',
                           usuario=None, roles=roles, procesos=procesos)


@app.route('/admin/usuarios/<int:id>/editar', methods=['GET', 'POST'])
@solo_admin
def usuario_editar(id):
    db = get_db()
    usuario = db.execute('SELECT * FROM usuario WHERE id=?', (id,)).fetchone()
    roles = db.execute('SELECT * FROM rol ORDER BY nivel').fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    if not usuario:
        flash('Usuario no encontrado.', 'danger')
        db.close()
        return redirect(url_for('usuarios_list'))
    if request.method == 'POST':
        nombre = request.form.get('nombre_completo', '').strip()
        email = request.form.get('email', '').strip()
        rol_id = request.form.get('rol_id')
        proceso_id = request.form.get('proceso_id') or None
        activo = 1 if request.form.get('activo') else 0
        debe_cambiar = 1 if request.form.get('debe_cambiar_password') else 0
        nueva_password = request.form.get('nueva_password', '').strip()

        if nueva_password:
            if len(nueva_password) < 8:
                flash('La nueva contraseña debe tener al menos 8 caracteres.', 'danger')
                db.close()
                return render_template('admin/usuario_form.html',
                                       usuario=usuario, roles=roles, procesos=procesos)
            db.execute('UPDATE usuario SET password_hash=? WHERE id=?',
                       (generate_password_hash(nueva_password), id))

        db.execute('''
            UPDATE usuario SET nombre_completo=?, email=?, rol_id=?,
            proceso_id=?, activo=?, debe_cambiar_password=? WHERE id=?
        ''', (nombre, email, rol_id, proceso_id, activo, debe_cambiar, id))
        db.commit()
        registrar_auditoria('EDITAR', 'Usuarios',
                            f'Editó usuario ID:{id} ({usuario["username"]})')
        flash('Usuario actualizado correctamente.', 'success')
        db.close()
        return redirect(url_for('usuarios_list'))
    db.close()
    return render_template('admin/usuario_form.html',
                           usuario=usuario, roles=roles, procesos=procesos)


@app.route('/admin/usuarios/<int:id>/toggle', methods=['POST'])
@solo_admin
def usuario_toggle(id):
    db = get_db()
    u = db.execute('SELECT * FROM usuario WHERE id=?', (id,)).fetchone()
    if u and u['id'] != session['user']['id']:
        nuevo = 0 if u['activo'] else 1
        db.execute('UPDATE usuario SET activo=? WHERE id=?', (nuevo, id))
        db.commit()
        estado = 'activado' if nuevo else 'desactivado'
        registrar_auditoria('TOGGLE', 'Usuarios',
                            f'Usuario {u["username"]} {estado}')
        flash(f'Usuario {u["username"]} {estado}.', 'success')
    else:
        flash('No puede desactivar su propia cuenta.', 'warning')
    db.close()
    return redirect(url_for('usuarios_list'))


@app.route('/admin/usuarios/<int:id>/eliminar', methods=['POST'])
@solo_admin
def usuario_eliminar(id):
    if id == session['user']['id']:
        flash('No puede eliminar su propia cuenta.', 'danger')
        return redirect(url_for('usuarios_list'))
    db = get_db()
    u = db.execute('SELECT username FROM usuario WHERE id=?', (id,)).fetchone()
    if u:
        db.execute('DELETE FROM usuario WHERE id=?', (id,))
        db.commit()
        registrar_auditoria('ELIMINAR', 'Usuarios', f'Eliminó usuario: {u["username"]}')
        flash(f'Usuario "{u["username"]}" eliminado.', 'warning')
    db.close()
    return redirect(url_for('usuarios_list'))


# ─── GESTIÓN DE ROLES ────────────────────────────────────────────────────────

@app.route('/admin/roles')
@solo_admin
def roles_list():
    db = get_db()
    roles = db.execute('''
        SELECT r.*, COUNT(u.id) as num_usuarios
        FROM rol r LEFT JOIN usuario u ON u.rol_id = r.id
        GROUP BY r.id ORDER BY r.nivel
    ''').fetchall()
    db.close()
    return render_template('admin/roles_list.html', roles=roles)


@app.route('/admin/roles/nuevo', methods=['GET', 'POST'])
@solo_admin
def rol_nuevo():
    if request.method == 'POST':
        db = get_db()
        db.execute('''
            INSERT INTO rol (nombre, nivel, descripcion,
                             puede_crear, puede_editar, puede_eliminar, puede_admin_usuarios)
            VALUES (?,?,?,?,?,?,?)
        ''', (request.form['nombre'], request.form['nivel'],
              request.form.get('descripcion', ''),
              1 if request.form.get('puede_crear') else 0,
              1 if request.form.get('puede_editar') else 0,
              1 if request.form.get('puede_eliminar') else 0,
              1 if request.form.get('puede_admin_usuarios') else 0))
        db.commit()
        registrar_auditoria('CREAR', 'Roles', f'Creó rol: {request.form["nombre"]}')
        flash('Rol creado correctamente.', 'success')
        db.close()
        return redirect(url_for('roles_list'))
    return render_template('admin/rol_form.html', rol=None)


@app.route('/admin/roles/<int:id>/editar', methods=['GET', 'POST'])
@solo_admin
def rol_editar(id):
    db = get_db()
    rol = db.execute('SELECT * FROM rol WHERE id=?', (id,)).fetchone()
    if request.method == 'POST':
        db.execute('''
            UPDATE rol SET nombre=?, nivel=?, descripcion=?,
            puede_crear=?, puede_editar=?, puede_eliminar=?, puede_admin_usuarios=?
            WHERE id=?
        ''', (request.form['nombre'], request.form['nivel'],
              request.form.get('descripcion', ''),
              1 if request.form.get('puede_crear') else 0,
              1 if request.form.get('puede_editar') else 0,
              1 if request.form.get('puede_eliminar') else 0,
              1 if request.form.get('puede_admin_usuarios') else 0, id))
        db.commit()
        registrar_auditoria('EDITAR', 'Roles', f'Editó rol ID:{id}')
        flash('Rol actualizado correctamente.', 'success')
        db.close()
        return redirect(url_for('roles_list'))
    db.close()
    return render_template('admin/rol_form.html', rol=rol)


@app.route('/admin/roles/<int:id>/eliminar', methods=['POST'])
@solo_admin
def rol_eliminar(id):
    db = get_db()
    r = db.execute('SELECT * FROM rol WHERE id=?', (id,)).fetchone()
    usuarios_con_rol = db.execute('SELECT COUNT(*) as c FROM usuario WHERE rol_id=?', (id,)).fetchone()
    if usuarios_con_rol['c'] > 0:
        flash('No puede eliminar un rol que tiene usuarios asignados.', 'danger')
    else:
        db.execute('DELETE FROM rol WHERE id=?', (id,))
        db.commit()
        registrar_auditoria('ELIMINAR', 'Roles', f'Eliminó rol: {r["nombre"]}')
        flash(f'Rol "{r["nombre"]}" eliminado.', 'warning')
    db.close()
    return redirect(url_for('roles_list'))


# ─── AUDITORÍA ───────────────────────────────────────────────────────────────

@app.route('/admin/auditoria')
@solo_admin
def auditoria():
    db = get_db()
    filtro_usuario = request.args.get('username', '')
    filtro_accion = request.args.get('accion', '')
    filtro_desde = request.args.get('desde', '')
    filtro_hasta = request.args.get('hasta', '')
    query = 'SELECT * FROM log_auditoria WHERE 1=1'
    params = []
    if filtro_usuario:
        query += ' AND username LIKE ?'
        params.append(f'%{filtro_usuario}%')
    if filtro_accion:
        query += ' AND accion=?'
        params.append(filtro_accion)
    if filtro_desde:
        query += ' AND fecha >= ?'
        params.append(filtro_desde)
    if filtro_hasta:
        query += ' AND fecha <= ?'
        params.append(filtro_hasta + ' 23:59:59')
    query += ' ORDER BY fecha DESC LIMIT 1000'
    logs = db.execute(query, params).fetchall()
    acciones = db.execute('SELECT DISTINCT accion FROM log_auditoria ORDER BY accion').fetchall()
    db.close()
    return render_template('admin/auditoria.html', logs=logs,
                           acciones=acciones, filtro_usuario=filtro_usuario,
                           filtro_accion=filtro_accion,
                           filtro_desde=filtro_desde, filtro_hasta=filtro_hasta)


# ─── DASHBOARD ──────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def dashboard():
    db = get_db()
    vid = get_vigencia_id(db)
    procesos = db.execute('''
        SELECT p.*, tp.nombre as tipo_nombre,
               COALESCE(scores.proc_score,0) as proc_score,
               COALESCE(scores.ind_score,0) as ind_score,
               COALESCE(scores.car_score,0) as car_score,
               COALESCE(scores.norm_score,0) as norm_score,
               COALESCE(scores.risk_score,0) as risk_score,
               COALESCE(
                 (COALESCE(scores.proc_score,0)+COALESCE(scores.ind_score,0)+
                  COALESCE(scores.car_score,0)+COALESCE(scores.norm_score,0)+
                  COALESCE(scores.risk_score,0))/5.0, 0) as ponderacion
        FROM proceso p
        LEFT JOIN tipo_proceso tp ON p.tipo_proceso_id = tp.id
        LEFT JOIN ponderacion_proceso scores ON p.id = scores.proceso_id
            AND scores.vigencia_id=?
        ORDER BY tp.orden, p.codigo
    ''', (vid,)).fetchall()
    total = len(procesos)
    ponds = [p['ponderacion'] for p in procesos]
    avance_global = round((sum(ponds) / total * 100) if total else 0, 1)
    resumen = {
        'total': total,
        'verde': sum(1 for p in procesos if semaforo(p['ponderacion']) == 'verde'),
        'amarillo': sum(1 for p in procesos if semaforo(p['ponderacion']) == 'amarillo'),
        'rojo': sum(1 for p in procesos if semaforo(p['ponderacion']) == 'rojo'),
        'avance_global': avance_global,
    }
    db.close()
    return render_template('dashboard.html', procesos=procesos, resumen=resumen)


# ─── TIPOS DE PROCESO ────────────────────────────────────────────────────────

@app.route('/tipos-proceso')
@login_required
def tipos_proceso():
    db = get_db()
    tipos = db.execute('SELECT * FROM tipo_proceso ORDER BY orden').fetchall()
    db.close()
    return render_template('tipos_proceso/list.html', tipos=tipos)


@app.route('/tipos-proceso/nuevo', methods=['GET', 'POST'])
@puede_editar
def tipo_proceso_nuevo():
    if request.method == 'POST':
        db = get_db()
        db.execute('INSERT INTO tipo_proceso (nombre, codigo, orden, descripcion) VALUES (?,?,?,?)',
                   (request.form['nombre'], request.form['codigo'],
                    request.form['orden'], request.form.get('descripcion', '')))
        db.commit()
        registrar_auditoria('CREAR', 'Tipos Proceso', request.form['nombre'])
        flash('Tipo de proceso creado exitosamente.', 'success')
        db.close()
        return redirect(url_for('tipos_proceso'))
    return render_template('tipos_proceso/form.html', tipo=None)


@app.route('/tipos-proceso/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def tipo_proceso_editar(id):
    db = get_db()
    tipo = db.execute('SELECT * FROM tipo_proceso WHERE id=?', (id,)).fetchone()
    if request.method == 'POST':
        db.execute('UPDATE tipo_proceso SET nombre=?, codigo=?, orden=?, descripcion=? WHERE id=?',
                   (request.form['nombre'], request.form['codigo'],
                    request.form['orden'], request.form.get('descripcion', ''), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Tipos Proceso', f'ID:{id}')
        flash('Tipo de proceso actualizado.', 'success')
        db.close()
        return redirect(url_for('tipos_proceso'))
    db.close()
    return render_template('tipos_proceso/form.html', tipo=tipo)


@app.route('/tipos-proceso/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def tipo_proceso_eliminar(id):
    db = get_db()
    db.execute('DELETE FROM tipo_proceso WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Tipos Proceso', f'ID:{id}')
    flash('Tipo de proceso eliminado.', 'warning')
    db.close()
    return redirect(url_for('tipos_proceso'))


# ─── ORGANIGRAMA ─────────────────────────────────────────────────────────────

@app.route('/organigrama')
@login_required
def organigrama():
    db = get_db()
    unidades = db.execute('''
        SELECT u.*, COUNT(p.id) as num_procesos
        FROM unidad_organigrama u
        LEFT JOIN proceso p ON p.unidad_id = u.id
        GROUP BY u.id ORDER BY u.nivel, u.nombre
    ''').fetchall()
    db.close()
    return render_template('organigrama/list.html', unidades=unidades)


@app.route('/organigrama/nuevo', methods=['GET', 'POST'])
@puede_editar
def organigrama_nuevo():
    if request.method == 'POST':
        db = get_db()
        db.execute('INSERT INTO unidad_organigrama (nombre, nivel, responsable, descripcion) VALUES (?,?,?,?)',
                   (request.form['nombre'], request.form['nivel'],
                    request.form.get('responsable', ''), request.form.get('descripcion', '')))
        db.commit()
        registrar_auditoria('CREAR', 'Organigrama', request.form['nombre'])
        flash('Unidad del organigrama creada.', 'success')
        db.close()
        return redirect(url_for('organigrama'))
    return render_template('organigrama/form.html', unidad=None)


@app.route('/organigrama/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def organigrama_editar(id):
    db = get_db()
    unidad = db.execute('SELECT * FROM unidad_organigrama WHERE id=?', (id,)).fetchone()
    if request.method == 'POST':
        db.execute('UPDATE unidad_organigrama SET nombre=?, nivel=?, responsable=?, descripcion=? WHERE id=?',
                   (request.form['nombre'], request.form['nivel'],
                    request.form.get('responsable', ''), request.form.get('descripcion', ''), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Organigrama', f'ID:{id}')
        flash('Unidad actualizada.', 'success')
        db.close()
        return redirect(url_for('organigrama'))
    db.close()
    return render_template('organigrama/form.html', unidad=unidad)


@app.route('/organigrama/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def organigrama_eliminar(id):
    db = get_db()
    db.execute('DELETE FROM unidad_organigrama WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Organigrama', f'ID:{id}')
    flash('Unidad eliminada.', 'warning')
    db.close()
    return redirect(url_for('organigrama'))


# ─── PROCESOS ────────────────────────────────────────────────────────────────

@app.route('/procesos')
@login_required
def procesos():
    db = get_db()
    vid = get_vigencia_id(db)
    lista = db.execute('''
        SELECT p.*, tp.nombre as tipo_nombre, u.nombre as unidad_nombre,
               COALESCE(s.proc_score,0) as proc_score,
               COALESCE(s.ind_score,0) as ind_score,
               COALESCE(s.car_score,0) as car_score,
               COALESCE(s.norm_score,0) as norm_score,
               COALESCE(s.risk_score,0) as risk_score,
               COALESCE((COALESCE(s.proc_score,0)+COALESCE(s.ind_score,0)+
                         COALESCE(s.car_score,0)+COALESCE(s.norm_score,0)+
                         COALESCE(s.risk_score,0))/5.0,0) as ponderacion
        FROM proceso p
        LEFT JOIN tipo_proceso tp ON p.tipo_proceso_id = tp.id
        LEFT JOIN unidad_organigrama u ON p.unidad_id = u.id
        LEFT JOIN ponderacion_proceso s ON p.id = s.proceso_id AND s.vigencia_id=?
        ORDER BY tp.orden, p.codigo
    ''', (vid,)).fetchall()
    db.close()
    return render_template('procesos/list.html', procesos=lista)


@app.route('/procesos/nuevo', methods=['GET', 'POST'])
@puede_editar
def proceso_nuevo():
    db = get_db()
    tipos = db.execute('SELECT * FROM tipo_proceso ORDER BY orden').fetchall()
    unidades = db.execute('SELECT * FROM unidad_organigrama ORDER BY nombre').fetchall()
    if request.method == 'POST':
        vid = get_vigencia_id(db)
        cur = db.execute(
            'INSERT INTO proceso (codigo, nombre, tipo_proceso_id, unidad_id, objetivo, alcance) VALUES (?,?,?,?,?,?)',
            (request.form['codigo'], request.form['nombre'],
             request.form['tipo_proceso_id'], request.form.get('unidad_id') or None,
             request.form.get('objetivo', ''), request.form.get('alcance', '')))
        pid = cur.lastrowid
        db.execute('INSERT OR IGNORE INTO ponderacion_proceso (proceso_id, vigencia_id) VALUES (?,?)', (pid, vid))
        db.commit()
        registrar_auditoria('CREAR', 'Procesos', request.form['nombre'])
        flash('Proceso creado exitosamente.', 'success')
        db.close()
        return redirect(url_for('procesos'))
    db.close()
    return render_template('procesos/form.html', proceso=None, tipos=tipos, unidades=unidades)


@app.route('/procesos/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def proceso_editar(id):
    db = get_db()
    proceso = db.execute('SELECT * FROM proceso WHERE id=?', (id,)).fetchone()
    tipos = db.execute('SELECT * FROM tipo_proceso ORDER BY orden').fetchall()
    unidades = db.execute('SELECT * FROM unidad_organigrama ORDER BY nombre').fetchall()
    if request.method == 'POST':
        db.execute(
            'UPDATE proceso SET codigo=?, nombre=?, tipo_proceso_id=?, unidad_id=?, objetivo=?, alcance=? WHERE id=?',
            (request.form['codigo'], request.form['nombre'],
             request.form['tipo_proceso_id'], request.form.get('unidad_id') or None,
             request.form.get('objetivo', ''), request.form.get('alcance', ''), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Procesos', f'{request.form["nombre"]} ID:{id}')
        flash('Proceso actualizado.', 'success')
        db.close()
        return redirect(url_for('procesos'))
    db.close()
    return render_template('procesos/form.html', proceso=proceso, tipos=tipos, unidades=unidades)


@app.route('/procesos/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def proceso_eliminar(id):
    db = get_db()
    p = db.execute('SELECT nombre FROM proceso WHERE id=?', (id,)).fetchone()
    db.execute('DELETE FROM proceso WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Procesos', p['nombre'] if p else f'ID:{id}')
    flash('Proceso eliminado.', 'warning')
    db.close()
    return redirect(url_for('procesos'))


@app.route('/procesos/<int:id>/detalle')
@login_required
def proceso_detalle(id):
    db = get_db()
    vid = get_vigencia_id(db)
    proceso = db.execute('''
        SELECT p.*, tp.nombre as tipo_nombre, u.nombre as unidad_nombre,
               COALESCE(s.proc_score,0) as proc_score,
               COALESCE(s.ind_score,0) as ind_score,
               COALESCE(s.car_score,0) as car_score,
               COALESCE(s.norm_score,0) as norm_score,
               COALESCE(s.risk_score,0) as risk_score,
               COALESCE((COALESCE(s.proc_score,0)+COALESCE(s.ind_score,0)+
                         COALESCE(s.car_score,0)+COALESCE(s.norm_score,0)+
                         COALESCE(s.risk_score,0))/5.0,0) as ponderacion
        FROM proceso p
        LEFT JOIN tipo_proceso tp ON p.tipo_proceso_id = tp.id
        LEFT JOIN unidad_organigrama u ON p.unidad_id = u.id
        LEFT JOIN ponderacion_proceso s ON p.id = s.proceso_id AND s.vigencia_id=?
        WHERE p.id=?
    ''', (vid, id)).fetchone()
    procedimientos = db.execute(
        'SELECT * FROM procedimiento WHERE proceso_id=? AND vigencia_id=? ORDER BY codigo', (id, vid)
    ).fetchall()
    indicadores = db.execute(
        'SELECT * FROM indicador WHERE proceso_id=? AND vigencia_id=? ORDER BY codigo', (id, vid)
    ).fetchall()
    caracterizaciones = db.execute(
        'SELECT * FROM caracterizacion WHERE proceso_id=? AND vigencia_id=?', (id, vid)
    ).fetchall()
    normogramas = db.execute(
        'SELECT * FROM normograma WHERE proceso_id=? AND vigencia_id=? ORDER BY fecha_expedicion DESC', (id, vid)
    ).fetchall()
    riesgos = db.execute(
        'SELECT * FROM riesgo WHERE proceso_id=? AND vigencia_id=? ORDER BY probabilidad DESC', (id, vid)
    ).fetchall()
    db.close()
    return render_template('procesos/detalle.html',
                           proceso=proceso, procedimientos=procedimientos,
                           indicadores=indicadores, caracterizaciones=caracterizaciones,
                           normogramas=normogramas, riesgos=riesgos)


@app.route('/procesos/<int:id>/ponderacion', methods=['GET', 'POST'])
@puede_editar
def proceso_ponderacion(id):
    db = get_db()
    vid = get_vigencia_id(db)
    proceso = db.execute('SELECT * FROM proceso WHERE id=?', (id,)).fetchone()
    pond = db.execute(
        'SELECT * FROM ponderacion_proceso WHERE proceso_id=? AND vigencia_id=?', (id, vid)
    ).fetchone()
    if request.method == 'POST':
        ps = float(request.form.get('proc_score', 0))
        is_ = float(request.form.get('ind_score', 0))
        cs = float(request.form.get('car_score', 0))
        ns = float(request.form.get('norm_score', 0))
        rs = float(request.form.get('risk_score', 0))
        db.execute('''
            INSERT INTO ponderacion_proceso (proceso_id, vigencia_id, proc_score, ind_score, car_score, norm_score, risk_score)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(proceso_id, vigencia_id) DO UPDATE SET
                proc_score=excluded.proc_score, ind_score=excluded.ind_score,
                car_score=excluded.car_score, norm_score=excluded.norm_score,
                risk_score=excluded.risk_score, fecha_actualizacion=CURRENT_TIMESTAMP
        ''', (id, vid, ps, is_, cs, ns, rs))
        db.commit()
        registrar_auditoria('EDITAR', 'Ponderación', f'Proceso ID:{id}')
        flash('Ponderación actualizada.', 'success')
        db.close()
        return redirect(url_for('proceso_detalle', id=id))
    db.close()
    return render_template('procesos/ponderacion.html', proceso=proceso, pond=pond)


# ─── PROCEDIMIENTOS ──────────────────────────────────────────────────────────

@app.route('/procedimientos')
@login_required
def procedimientos():
    db = get_db()
    vid = get_vigencia_id(db)
    filtro_proceso = request.args.get('proceso_id', '')
    query = '''
        SELECT pr.*, p.nombre as proceso_nombre, p.codigo as proceso_codigo
        FROM procedimiento pr JOIN proceso p ON pr.proceso_id = p.id
        WHERE pr.vigencia_id=?
    '''
    params = [vid]
    if filtro_proceso:
        query += ' AND pr.proceso_id=?'
        params.append(filtro_proceso)
    query += ' ORDER BY p.codigo, pr.codigo'
    lista = db.execute(query, params).fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    db.close()
    return render_template('procedimientos/list.html', procedimientos=lista,
                           procesos=procesos, filtro_proceso=filtro_proceso)


@app.route('/procedimientos/nuevo', methods=['GET', 'POST'])
@puede_editar
def procedimiento_nuevo():
    db = get_db()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        vid = get_vigencia_id(db)
        db.execute(
            'INSERT INTO procedimiento (proceso_id, codigo, nombre, estado, publicado_onedrive, descripcion, responsable, version, vigencia_id) VALUES (?,?,?,?,?,?,?,?,?)',
            (request.form['proceso_id'], request.form['codigo'], request.form['nombre'],
             request.form['estado'], request.form.get('publicado_onedrive', 'NO'),
             request.form.get('descripcion', ''), request.form.get('responsable', ''),
             request.form.get('version', '1.0'), vid))
        db.commit()
        registrar_auditoria('CREAR', 'Procedimientos', request.form['nombre'])
        flash('Procedimiento creado exitosamente.', 'success')
        db.close()
        return redirect(url_for('procedimientos'))
    pid = request.args.get('proceso_id')
    db.close()
    return render_template('procedimientos/form.html', proc=None,
                           procesos=procesos, selected_proceso=pid)


@app.route('/procedimientos/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def procedimiento_editar(id):
    db = get_db()
    proc = db.execute('SELECT * FROM procedimiento WHERE id=?', (id,)).fetchone()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        db.execute(
            'UPDATE procedimiento SET proceso_id=?, codigo=?, nombre=?, estado=?, publicado_onedrive=?, descripcion=?, responsable=?, version=? WHERE id=?',
            (request.form['proceso_id'], request.form['codigo'], request.form['nombre'],
             request.form['estado'], request.form.get('publicado_onedrive', 'NO'),
             request.form.get('descripcion', ''), request.form.get('responsable', ''),
             request.form.get('version', '1.0'), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Procedimientos', f'{request.form["nombre"]} ID:{id}')
        flash('Procedimiento actualizado.', 'success')
        db.close()
        return redirect(url_for('procedimientos'))
    db.close()
    return render_template('procedimientos/form.html', proc=proc,
                           procesos=procesos, selected_proceso=None)


@app.route('/procedimientos/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def procedimiento_eliminar(id):
    db = get_db()
    pr = db.execute('SELECT nombre FROM procedimiento WHERE id=?', (id,)).fetchone()
    db.execute('DELETE FROM procedimiento WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Procedimientos', pr['nombre'] if pr else f'ID:{id}')
    flash('Procedimiento eliminado.', 'warning')
    db.close()
    return redirect(url_for('procedimientos'))


# ─── INDICADORES ─────────────────────────────────────────────────────────────

@app.route('/indicadores')
@login_required
def indicadores():
    db = get_db()
    vid = get_vigencia_id(db)
    filtro_proceso = request.args.get('proceso_id', '')
    query = '''
        SELECT i.*, p.nombre as proceso_nombre, p.codigo as proceso_codigo
        FROM indicador i JOIN proceso p ON i.proceso_id = p.id
        WHERE i.vigencia_id=?
    '''
    params = [vid]
    if filtro_proceso:
        query += ' AND i.proceso_id=?'
        params.append(filtro_proceso)
    query += ' ORDER BY p.codigo, i.codigo'
    lista = db.execute(query, params).fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    db.close()
    return render_template('indicadores/list.html', indicadores=lista,
                           procesos=procesos, filtro_proceso=filtro_proceso)


@app.route('/indicadores/nuevo', methods=['GET', 'POST'])
@puede_editar
def indicador_nuevo():
    db = get_db()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        vid = get_vigencia_id(db)
        db.execute(
            'INSERT INTO indicador (proceso_id, codigo, nombre, estado, publicado_onedrive, formula, meta, unidad_medida, frecuencia, responsable, vigencia_id) VALUES (?,?,?,?,?,?,?,?,?,?,?)',
            (request.form['proceso_id'], request.form['codigo'], request.form['nombre'],
             request.form['estado'], request.form.get('publicado_onedrive', 'NO'),
             request.form.get('formula', ''), request.form.get('meta', ''),
             request.form.get('unidad_medida', ''), request.form.get('frecuencia', ''),
             request.form.get('responsable', ''), vid))
        db.commit()
        registrar_auditoria('CREAR', 'Indicadores', request.form['nombre'])
        flash('Indicador creado exitosamente.', 'success')
        db.close()
        return redirect(url_for('indicadores'))
    pid = request.args.get('proceso_id')
    db.close()
    return render_template('indicadores/form.html', ind=None,
                           procesos=procesos, selected_proceso=pid)


@app.route('/indicadores/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def indicador_editar(id):
    db = get_db()
    ind = db.execute('SELECT * FROM indicador WHERE id=?', (id,)).fetchone()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        db.execute(
            'UPDATE indicador SET proceso_id=?, codigo=?, nombre=?, estado=?, publicado_onedrive=?, formula=?, meta=?, unidad_medida=?, frecuencia=?, responsable=? WHERE id=?',
            (request.form['proceso_id'], request.form['codigo'], request.form['nombre'],
             request.form['estado'], request.form.get('publicado_onedrive', 'NO'),
             request.form.get('formula', ''), request.form.get('meta', ''),
             request.form.get('unidad_medida', ''), request.form.get('frecuencia', ''),
             request.form.get('responsable', ''), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Indicadores', f'ID:{id}')
        flash('Indicador actualizado.', 'success')
        db.close()
        return redirect(url_for('indicadores'))
    db.close()
    return render_template('indicadores/form.html', ind=ind,
                           procesos=procesos, selected_proceso=None)


@app.route('/indicadores/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def indicador_eliminar(id):
    db = get_db()
    i = db.execute('SELECT nombre FROM indicador WHERE id=?', (id,)).fetchone()
    db.execute('DELETE FROM indicador WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Indicadores', i['nombre'] if i else f'ID:{id}')
    flash('Indicador eliminado.', 'warning')
    db.close()
    return redirect(url_for('indicadores'))


# ─── CARACTERIZACIÓN ─────────────────────────────────────────────────────────

@app.route('/caracterizacion')
@login_required
def caracterizacion():
    db = get_db()
    vid = get_vigencia_id(db)
    filtro_proceso = request.args.get('proceso_id', '')
    query = '''
        SELECT c.*, p.nombre as proceso_nombre, p.codigo as proceso_codigo
        FROM caracterizacion c JOIN proceso p ON c.proceso_id = p.id
        WHERE c.vigencia_id=?
    '''
    params = [vid]
    if filtro_proceso:
        query += ' AND c.proceso_id=?'
        params.append(filtro_proceso)
    query += ' ORDER BY p.codigo'
    lista = db.execute(query, params).fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    db.close()
    return render_template('caracterizacion/list.html', caracterizaciones=lista,
                           procesos=procesos, filtro_proceso=filtro_proceso)


@app.route('/caracterizacion/nuevo', methods=['GET', 'POST'])
@puede_editar
def caracterizacion_nuevo():
    db = get_db()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        vid = get_vigencia_id(db)
        db.execute('''
            INSERT INTO caracterizacion
            (proceso_id, version, objetivo, alcance, proveedor, entradas, actividades,
             salidas, cliente, recursos_humanos, recursos_tecnologicos, recursos_fisicos,
             normatividad, indicadores_ref, riesgos_ref, estado, vigencia_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (request.form['proceso_id'], request.form.get('version', '1.0'),
              request.form.get('objetivo', ''), request.form.get('alcance', ''),
              request.form.get('proveedor', ''), request.form.get('entradas', ''),
              request.form.get('actividades', ''), request.form.get('salidas', ''),
              request.form.get('cliente', ''), request.form.get('recursos_humanos', ''),
              request.form.get('recursos_tecnologicos', ''), request.form.get('recursos_fisicos', ''),
              request.form.get('normatividad', ''), request.form.get('indicadores_ref', ''),
              request.form.get('riesgos_ref', ''), request.form.get('estado', 'BORRADOR'), vid))
        db.commit()
        registrar_auditoria('CREAR', 'Caracterización', f'Proceso ID:{request.form["proceso_id"]}')
        flash('Caracterización creada exitosamente.', 'success')
        db.close()
        return redirect(url_for('caracterizacion'))
    pid = request.args.get('proceso_id')
    db.close()
    return render_template('caracterizacion/form.html', car=None,
                           procesos=procesos, selected_proceso=pid)


@app.route('/caracterizacion/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def caracterizacion_editar(id):
    db = get_db()
    car = db.execute('SELECT * FROM caracterizacion WHERE id=?', (id,)).fetchone()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        db.execute('''
            UPDATE caracterizacion SET proceso_id=?, version=?, objetivo=?, alcance=?,
            proveedor=?, entradas=?, actividades=?, salidas=?, cliente=?,
            recursos_humanos=?, recursos_tecnologicos=?, recursos_fisicos=?,
            normatividad=?, indicadores_ref=?, riesgos_ref=?, estado=?,
            fecha_actualizacion=CURRENT_TIMESTAMP WHERE id=?
        ''', (request.form['proceso_id'], request.form.get('version', '1.0'),
              request.form.get('objetivo', ''), request.form.get('alcance', ''),
              request.form.get('proveedor', ''), request.form.get('entradas', ''),
              request.form.get('actividades', ''), request.form.get('salidas', ''),
              request.form.get('cliente', ''), request.form.get('recursos_humanos', ''),
              request.form.get('recursos_tecnologicos', ''), request.form.get('recursos_fisicos', ''),
              request.form.get('normatividad', ''), request.form.get('indicadores_ref', ''),
              request.form.get('riesgos_ref', ''), request.form.get('estado', 'BORRADOR'), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Caracterización', f'ID:{id}')
        flash('Caracterización actualizada.', 'success')
        db.close()
        return redirect(url_for('caracterizacion'))
    db.close()
    return render_template('caracterizacion/form.html', car=car,
                           procesos=procesos, selected_proceso=None)


@app.route('/caracterizacion/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def caracterizacion_eliminar(id):
    db = get_db()
    db.execute('DELETE FROM caracterizacion WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Caracterización', f'ID:{id}')
    flash('Caracterización eliminada.', 'warning')
    db.close()
    return redirect(url_for('caracterizacion'))


# ─── NORMOGRAMA ──────────────────────────────────────────────────────────────

@app.route('/normograma')
@login_required
def normograma():
    db = get_db()
    vid = get_vigencia_id(db)
    filtro_proceso = request.args.get('proceso_id', '')
    query = '''
        SELECT n.*, p.nombre as proceso_nombre
        FROM normograma n JOIN proceso p ON n.proceso_id = p.id
        WHERE n.vigencia_id=?
    '''
    params = [vid]
    if filtro_proceso:
        query += ' AND n.proceso_id=?'
        params.append(filtro_proceso)
    query += ' ORDER BY n.fecha_expedicion DESC'
    lista = db.execute(query, params).fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    db.close()
    return render_template('normograma/list.html', normas=lista,
                           procesos=procesos, filtro_proceso=filtro_proceso)


@app.route('/normograma/nuevo', methods=['GET', 'POST'])
@puede_editar
def normograma_nuevo():
    db = get_db()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        vid = get_vigencia_id(db)
        db.execute('''
            INSERT INTO normograma
            (proceso_id, tipo_norma, numero, año, entidad_expide, titulo, descripcion,
             fecha_expedicion, vigente, enlace, vigencia_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)
        ''', (request.form['proceso_id'], request.form['tipo_norma'],
              request.form.get('numero', ''), request.form.get('año', ''),
              request.form.get('entidad_expide', ''), request.form.get('titulo', ''),
              request.form.get('descripcion', ''), request.form.get('fecha_expedicion') or None,
              1 if request.form.get('vigente') else 0, request.form.get('enlace', ''), vid))
        db.commit()
        registrar_auditoria('CREAR', 'Normograma', request.form.get('titulo', ''))
        flash('Norma registrada exitosamente.', 'success')
        db.close()
        return redirect(url_for('normograma'))
    pid = request.args.get('proceso_id')
    db.close()
    return render_template('normograma/form.html', norma=None,
                           procesos=procesos, selected_proceso=pid)


@app.route('/normograma/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def normograma_editar(id):
    db = get_db()
    norma = db.execute('SELECT * FROM normograma WHERE id=?', (id,)).fetchone()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        db.execute('''
            UPDATE normograma SET proceso_id=?, tipo_norma=?, numero=?, año=?,
            entidad_expide=?, titulo=?, descripcion=?, fecha_expedicion=?,
            vigente=?, enlace=? WHERE id=?
        ''', (request.form['proceso_id'], request.form['tipo_norma'],
              request.form.get('numero', ''), request.form.get('año', ''),
              request.form.get('entidad_expide', ''), request.form.get('titulo', ''),
              request.form.get('descripcion', ''), request.form.get('fecha_expedicion') or None,
              1 if request.form.get('vigente') else 0, request.form.get('enlace', ''), id))
        db.commit()
        registrar_auditoria('EDITAR', 'Normograma', f'ID:{id}')
        flash('Norma actualizada.', 'success')
        db.close()
        return redirect(url_for('normograma'))
    db.close()
    return render_template('normograma/form.html', norma=norma,
                           procesos=procesos, selected_proceso=None)


@app.route('/normograma/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def normograma_eliminar(id):
    db = get_db()
    n = db.execute('SELECT titulo FROM normograma WHERE id=?', (id,)).fetchone()
    db.execute('DELETE FROM normograma WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Normograma', n['titulo'] if n else f'ID:{id}')
    flash('Norma eliminada.', 'warning')
    db.close()
    return redirect(url_for('normograma'))


# ─── MAPA DE RIESGOS ─────────────────────────────────────────────────────────

NIVELES = {
    (1,1):('Bajo','#28a745'),(1,2):('Bajo','#28a745'),(1,3):('Moderado','#ffc107'),
    (1,4):('Alto','#fd7e14'),(1,5):('Extremo','#dc3545'),
    (2,1):('Bajo','#28a745'),(2,2):('Moderado','#ffc107'),(2,3):('Alto','#fd7e14'),
    (2,4):('Alto','#fd7e14'),(2,5):('Extremo','#dc3545'),
    (3,1):('Moderado','#ffc107'),(3,2):('Alto','#fd7e14'),(3,3):('Alto','#fd7e14'),
    (3,4):('Extremo','#dc3545'),(3,5):('Extremo','#dc3545'),
    (4,1):('Alto','#fd7e14'),(4,2):('Alto','#fd7e14'),(4,3):('Extremo','#dc3545'),
    (4,4):('Extremo','#dc3545'),(4,5):('Extremo','#dc3545'),
    (5,1):('Extremo','#dc3545'),(5,2):('Extremo','#dc3545'),(5,3):('Extremo','#dc3545'),
    (5,4):('Extremo','#dc3545'),(5,5):('Extremo','#dc3545'),
}


@app.route('/riesgos')
@login_required
def riesgos():
    db = get_db()
    vid = get_vigencia_id(db)
    filtro_proceso = request.args.get('proceso_id', '')
    query = '''
        SELECT r.*, p.nombre as proceso_nombre
        FROM riesgo r JOIN proceso p ON r.proceso_id = p.id
        WHERE r.vigencia_id=?
    '''
    params = [vid]
    if filtro_proceso:
        query += ' AND r.proceso_id=?'
        params.append(filtro_proceso)
    query += ' ORDER BY (r.probabilidad * r.impacto) DESC'
    lista = db.execute(query, params).fetchall()
    procesos = db.execute('SELECT id, nombre FROM proceso ORDER BY nombre').fetchall()
    db.close()
    return render_template('riesgos/list.html', riesgos=lista,
                           procesos=procesos, filtro_proceso=filtro_proceso, niveles=NIVELES)


@app.route('/riesgos/nuevo', methods=['GET', 'POST'])
@puede_editar
def riesgo_nuevo():
    db = get_db()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        prob = int(request.form.get('probabilidad', 1))
        imp = int(request.form.get('impacto', 1))
        nivel = NIVELES.get((prob, imp), ('Bajo', '#28a745'))[0]
        vid = get_vigencia_id(db)
        db.execute('''
            INSERT INTO riesgo
            (proceso_id, codigo, nombre, descripcion, tipo_riesgo, causa,
             consecuencia, probabilidad, impacto, nivel_riesgo, control_existente,
             accion_mitigacion, responsable, fecha_identificacion, vigencia_id)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''', (request.form['proceso_id'], request.form.get('codigo', ''),
              request.form['nombre'], request.form.get('descripcion', ''),
              request.form.get('tipo_riesgo', ''), request.form.get('causa', ''),
              request.form.get('consecuencia', ''), prob, imp, nivel,
              request.form.get('control_existente', ''),
              request.form.get('accion_mitigacion', ''),
              request.form.get('responsable', ''),
              request.form.get('fecha_identificacion') or None, vid))
        db.commit()
        registrar_auditoria('CREAR', 'Riesgos', request.form['nombre'])
        flash('Riesgo registrado exitosamente.', 'success')
        db.close()
        return redirect(url_for('riesgos'))
    pid = request.args.get('proceso_id')
    db.close()
    return render_template('riesgos/form.html', riesgo=None,
                           procesos=procesos, selected_proceso=pid)


@app.route('/riesgos/<int:id>/editar', methods=['GET', 'POST'])
@puede_editar
def riesgo_editar(id):
    db = get_db()
    riesgo = db.execute('SELECT * FROM riesgo WHERE id=?', (id,)).fetchone()
    procesos = db.execute('SELECT id, nombre, codigo FROM proceso ORDER BY nombre').fetchall()
    if request.method == 'POST':
        prob = int(request.form.get('probabilidad', 1))
        imp = int(request.form.get('impacto', 1))
        nivel = NIVELES.get((prob, imp), ('Bajo', '#28a745'))[0]
        db.execute('''
            UPDATE riesgo SET proceso_id=?, codigo=?, nombre=?, descripcion=?, tipo_riesgo=?,
            causa=?, consecuencia=?, probabilidad=?, impacto=?, nivel_riesgo=?,
            control_existente=?, accion_mitigacion=?, responsable=?, fecha_identificacion=?
            WHERE id=?
        ''', (request.form['proceso_id'], request.form.get('codigo', ''),
              request.form['nombre'], request.form.get('descripcion', ''),
              request.form.get('tipo_riesgo', ''), request.form.get('causa', ''),
              request.form.get('consecuencia', ''), prob, imp, nivel,
              request.form.get('control_existente', ''),
              request.form.get('accion_mitigacion', ''),
              request.form.get('responsable', ''),
              request.form.get('fecha_identificacion') or None, id))
        db.commit()
        registrar_auditoria('EDITAR', 'Riesgos', f'ID:{id}')
        flash('Riesgo actualizado.', 'success')
        db.close()
        return redirect(url_for('riesgos'))
    db.close()
    return render_template('riesgos/form.html', riesgo=riesgo,
                           procesos=procesos, selected_proceso=None)


@app.route('/riesgos/<int:id>/eliminar', methods=['POST'])
@puede_eliminar
def riesgo_eliminar(id):
    db = get_db()
    r = db.execute('SELECT nombre FROM riesgo WHERE id=?', (id,)).fetchone()
    db.execute('DELETE FROM riesgo WHERE id=?', (id,))
    db.commit()
    registrar_auditoria('ELIMINAR', 'Riesgos', r['nombre'] if r else f'ID:{id}')
    flash('Riesgo eliminado.', 'warning')
    db.close()
    return redirect(url_for('riesgos'))


# ─── BÚSQUEDA GLOBAL ─────────────────────────────────────────────────────────

@app.route('/buscar')
@login_required
def buscar():
    q = request.args.get('q', '').strip()
    if not q or len(q) < 2:
        return render_template('buscar.html', resultados={}, q=q, total=0)
    db = get_db()
    vid = get_vigencia_id(db)
    like = f'%{q}%'
    procs = db.execute(
        'SELECT id, codigo, nombre FROM proceso WHERE nombre LIKE ? OR codigo LIKE ? LIMIT 10',
        (like, like)).fetchall()
    procs_det = db.execute(
        '''SELECT pr.id, pr.codigo, pr.nombre, p.nombre as proceso_nombre
           FROM procedimiento pr JOIN proceso p ON pr.proceso_id = p.id
           WHERE pr.vigencia_id=? AND (pr.nombre LIKE ? OR pr.codigo LIKE ?) LIMIT 10''',
        (vid, like, like)).fetchall()
    normas = db.execute(
        '''SELECT n.id, n.titulo, n.tipo_norma, n.numero, p.nombre as proceso_nombre
           FROM normograma n JOIN proceso p ON n.proceso_id = p.id
           WHERE n.vigencia_id=? AND (n.titulo LIKE ? OR n.numero LIKE ? OR n.descripcion LIKE ?) LIMIT 10''',
        (vid, like, like, like)).fetchall()
    riesgos_r = db.execute(
        '''SELECT r.id, r.codigo, r.nombre, r.nivel_riesgo, p.nombre as proceso_nombre
           FROM riesgo r JOIN proceso p ON r.proceso_id = p.id
           WHERE r.vigencia_id=? AND (r.nombre LIKE ? OR r.descripcion LIKE ?) LIMIT 10''',
        (vid, like, like)).fetchall()
    indicad = db.execute(
        '''SELECT i.id, i.codigo, i.nombre, p.nombre as proceso_nombre
           FROM indicador i JOIN proceso p ON i.proceso_id = p.id
           WHERE i.vigencia_id=? AND (i.nombre LIKE ? OR i.codigo LIKE ?) LIMIT 10''',
        (vid, like, like)).fetchall()
    db.close()
    resultados = {
        'procesos': procs,
        'procedimientos': procs_det,
        'indicadores': indicad,
        'normas': normas,
        'riesgos': riesgos_r,
    }
    total = sum(len(v) for v in resultados.values())
    return render_template('buscar.html', resultados=resultados, q=q, total=total)


# ─── EXPORTACIÓN ─────────────────────────────────────────────────────────────

def _excel_header_style():
    font = Font(bold=True, color='FFFFFF', size=10)
    fill = PatternFill('solid', fgColor='1A2744')
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    return font, fill, align


@app.route('/exportar/dashboard')
@login_required
def exportar_dashboard():
    db = get_db()
    vid = get_vigencia_id(db)
    rows = db.execute('''
        SELECT p.codigo, p.nombre, tp.nombre as tipo,
               COALESCE(s.proc_score,0)*100 as proc,
               COALESCE(s.ind_score,0)*100 as ind,
               COALESCE(s.car_score,0)*100 as car,
               COALESCE(s.norm_score,0)*100 as norm,
               COALESCE(s.risk_score,0)*100 as risk,
               COALESCE((COALESCE(s.proc_score,0)+COALESCE(s.ind_score,0)+
                         COALESCE(s.car_score,0)+COALESCE(s.norm_score,0)+
                         COALESCE(s.risk_score,0))/5.0,0)*100 as ponderacion
        FROM proceso p
        LEFT JOIN tipo_proceso tp ON p.tipo_proceso_id = tp.id
        LEFT JOIN ponderacion_proceso s ON p.id = s.proceso_id AND s.vigencia_id=?
        ORDER BY tp.orden, p.codigo
    ''', (vid,)).fetchall()
    db.close()

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tablero GICA'
        font_h, fill_h, align_h = _excel_header_style()
        headers = ['Código', 'Proceso', 'Tipo', 'Proced. %', 'Indicad. %',
                   'Caracteriz. %', 'Normograma %', 'M. Riesgos %', 'Total %', 'Estado']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font_h
            cell.fill = fill_h
            cell.alignment = align_h
        colores_sem = {'Verde': 'C6EFCE', 'Amarillo': 'FFEB9C', 'Rojo': 'FFC7CE'}
        for row_i, p in enumerate(rows, 2):
            pond = round(p['ponderacion'], 1)
            estado = 'Verde' if pond >= 90 else ('Amarillo' if pond >= 70 else 'Rojo')
            data = [p['codigo'], p['nombre'], p['tipo'],
                    round(p['proc'], 1), round(p['ind'], 1), round(p['car'], 1),
                    round(p['norm'], 1), round(p['risk'], 1), pond, estado]
            for col_i, val in enumerate(data, 1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                if estado in colores_sem:
                    cell.fill = PatternFill('solid', fgColor=colores_sem[estado])
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 18
        for col_letter in ['A', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col_letter].width = 14
        ws.freeze_panes = 'A2'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'Tablero_GICA_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        output = io.StringIO()
        writer = csv_module.writer(output)
        writer.writerow(['Código', 'Proceso', 'Tipo', 'Proced.%', 'Indicad.%',
                         'Caracteriz.%', 'Normograma%', 'M.Riesgos%', 'Total%', 'Estado'])
        for p in rows:
            pond = round(p['ponderacion'], 1)
            estado = 'Verde' if pond >= 90 else ('Amarillo' if pond >= 70 else 'Rojo')
            writer.writerow([p['codigo'], p['nombre'], p['tipo'],
                             round(p['proc'], 1), round(p['ind'], 1), round(p['car'], 1),
                             round(p['norm'], 1), round(p['risk'], 1), pond, estado])
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename=Tablero_GICA_{datetime.now().strftime("%Y%m%d")}.csv'
        return resp


@app.route('/exportar/riesgos')
@login_required
def exportar_riesgos():
    db = get_db()
    vid = get_vigencia_id(db)
    rows = db.execute('''
        SELECT r.codigo, r.nombre, r.descripcion, r.tipo_riesgo, r.causa,
               r.consecuencia, r.probabilidad, r.impacto, r.nivel_riesgo,
               r.control_existente, r.accion_mitigacion, r.responsable,
               r.fecha_identificacion, p.nombre as proceso_nombre
        FROM riesgo r JOIN proceso p ON r.proceso_id = p.id
        WHERE r.vigencia_id=?
        ORDER BY (r.probabilidad * r.impacto) DESC
    ''', (vid,)).fetchall()
    db.close()

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Mapa de Riesgos'
        font_h, fill_h, align_h = _excel_header_style()
        headers = ['Código', 'Riesgo', 'Proceso', 'Tipo', 'Causa', 'Consecuencia',
                   'Probabilidad', 'Impacto', 'Nivel', 'Control Existente',
                   'Acción Mitigación', 'Responsable', 'Fecha Identificación']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font_h
            cell.fill = fill_h
            cell.alignment = align_h
        nivel_colors = {'Extremo': 'FFC7CE', 'Alto': 'FFEB9C', 'Moderado': 'DDEBF7', 'Bajo': 'C6EFCE'}
        for row_i, r in enumerate(rows, 2):
            data = [r['codigo'], r['nombre'], r['proceso_nombre'], r['tipo_riesgo'],
                    r['causa'], r['consecuencia'], r['probabilidad'], r['impacto'],
                    r['nivel_riesgo'], r['control_existente'], r['accion_mitigacion'],
                    r['responsable'], r['fecha_identificacion']]
            for col_i, val in enumerate(data, 1):
                cell = ws.cell(row=row_i, column=col_i, value=val)
                nivel = r['nivel_riesgo'] or 'Bajo'
                if nivel in nivel_colors:
                    cell.fill = PatternFill('solid', fgColor=nivel_colors[nivel])
        for col_letter, width in [('A', 10), ('B', 35), ('C', 30), ('D', 15),
                                   ('E', 30), ('F', 30), ('G', 12), ('H', 10),
                                   ('I', 12), ('J', 35), ('K', 35), ('L', 20), ('M', 15)]:
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = 'A2'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'Riesgos_GICA_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        output = io.StringIO()
        writer = csv_module.writer(output)
        writer.writerow(['Código', 'Riesgo', 'Proceso', 'Tipo', 'Causa', 'Consecuencia',
                         'Probabilidad', 'Impacto', 'Nivel', 'Control', 'Mitigación', 'Responsable', 'Fecha'])
        for r in rows:
            writer.writerow([r['codigo'], r['nombre'], r['proceso_nombre'], r['tipo_riesgo'],
                             r['causa'], r['consecuencia'], r['probabilidad'], r['impacto'],
                             r['nivel_riesgo'], r['control_existente'], r['accion_mitigacion'],
                             r['responsable'], r['fecha_identificacion']])
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename=Riesgos_GICA_{datetime.now().strftime("%Y%m%d")}.csv'
        return resp


@app.route('/exportar/auditoria')
@solo_admin
def exportar_auditoria():
    db = get_db()
    rows = db.execute(
        'SELECT * FROM log_auditoria ORDER BY fecha DESC LIMIT 5000'
    ).fetchall()
    db.close()

    if HAS_OPENPYXL:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Log Auditoría'
        font_h, fill_h, align_h = _excel_header_style()
        headers = ['ID', 'Usuario', 'Acción', 'Módulo', 'Detalle', 'IP', 'Fecha y Hora']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = font_h
            cell.fill = fill_h
            cell.alignment = align_h
        for row_i, log in enumerate(rows, 2):
            ws.append([log['id'], log['username'], log['accion'], log['modulo'],
                       log['detalle'], log['ip'], log['fecha']])
        for col_letter, width in [('A', 8), ('B', 18), ('C', 18), ('D', 18), ('E', 50), ('F', 15), ('G', 20)]:
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = 'A2'
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'Auditoria_GICA_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(output, download_name=filename, as_attachment=True,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        output = io.StringIO()
        writer = csv_module.writer(output)
        writer.writerow(['ID', 'Usuario', 'Acción', 'Módulo', 'Detalle', 'IP', 'Fecha'])
        for log in rows:
            writer.writerow([log['id'], log['username'], log['accion'], log['modulo'],
                             log['detalle'], log['ip'], log['fecha']])
        resp = make_response(output.getvalue())
        resp.headers['Content-Type'] = 'text/csv; charset=utf-8'
        resp.headers['Content-Disposition'] = f'attachment; filename=Auditoria_GICA_{datetime.now().strftime("%Y%m%d")}.csv'
        return resp


# ─── API JSON ────────────────────────────────────────────────────────────────

@app.route('/api/ponderacion-data')
@login_required
def api_ponderacion_data():
    db = get_db()
    vid = get_vigencia_id(db)
    data = db.execute('''
        SELECT p.nombre,
               COALESCE((COALESCE(s.proc_score,0)+COALESCE(s.ind_score,0)+
                         COALESCE(s.car_score,0)+COALESCE(s.norm_score,0)+
                         COALESCE(s.risk_score,0))/5.0,0) as ponderacion
        FROM proceso p
        LEFT JOIN ponderacion_proceso s ON p.id = s.proceso_id AND s.vigencia_id=?
        ORDER BY ponderacion DESC
    ''', (vid,)).fetchall()
    db.close()
    return jsonify([{'nombre': r['nombre'], 'valor': round(r['ponderacion'] * 100, 1)} for r in data])


# ─── VIGENCIAS ───────────────────────────────────────────────────────────────

@app.route('/admin/vigencias')
@solo_admin
def vigencias_list():
    db = get_db()
    vigencias = db.execute('''
        SELECT v.*,
               (SELECT COUNT(*) FROM ponderacion_proceso WHERE vigencia_id=v.id) as num_procesos,
               (SELECT COUNT(*) FROM procedimiento WHERE vigencia_id=v.id) as num_procedimientos,
               (SELECT COUNT(*) FROM indicador WHERE vigencia_id=v.id) as num_indicadores,
               (SELECT COUNT(*) FROM riesgo WHERE vigencia_id=v.id) as num_riesgos
        FROM vigencia v
        ORDER BY v.año DESC
    ''').fetchall()
    db.close()
    return render_template('vigencias/list.html', vigencias=vigencias)


@app.route('/admin/vigencias/nueva', methods=['GET', 'POST'])
@solo_admin
def vigencia_nueva():
    if request.method == 'POST':
        try:
            año = int(request.form['año'])
        except (ValueError, KeyError):
            flash('El año debe ser un número válido.', 'danger')
            return render_template('vigencias/form.html', vigencia=None)
        nombre = request.form.get('nombre', '').strip() or f'GICA {año}'
        descripcion = request.form.get('descripcion', '')
        fecha_inicio = request.form.get('fecha_inicio') or None
        db = get_db()
        if db.execute('SELECT id FROM vigencia WHERE año=?', (año,)).fetchone():
            flash(f'Ya existe una vigencia para el año {año}.', 'danger')
            db.close()
            return render_template('vigencias/form.html', vigencia=None)
        db.execute('''
            INSERT INTO vigencia (año, nombre, descripcion, activa, fecha_inicio)
            VALUES (?,?,?,0,?)
        ''', (año, nombre, descripcion, fecha_inicio))
        db.commit()
        registrar_auditoria('CREAR', 'Vigencias', f'Vigencia {nombre}')
        flash(f'Vigencia "{nombre}" creada exitosamente.', 'success')
        db.close()
        return redirect(url_for('vigencias_list'))
    return render_template('vigencias/form.html', vigencia=None)


@app.route('/admin/vigencias/<int:id>/editar', methods=['GET', 'POST'])
@solo_admin
def vigencia_editar(id):
    db = get_db()
    v = db.execute('SELECT * FROM vigencia WHERE id=?', (id,)).fetchone()
    if not v:
        flash('Vigencia no encontrada.', 'danger')
        db.close()
        return redirect(url_for('vigencias_list'))
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip() or v['nombre']
        descripcion = request.form.get('descripcion', '')
        fecha_inicio = request.form.get('fecha_inicio') or None
        fecha_cierre = request.form.get('fecha_cierre') or None
        db.execute('''
            UPDATE vigencia SET nombre=?, descripcion=?, fecha_inicio=?, fecha_cierre=?
            WHERE id=?
        ''', (nombre, descripcion, fecha_inicio, fecha_cierre, id))
        db.commit()
        registrar_auditoria('EDITAR', 'Vigencias', f'Vigencia ID:{id}')
        flash('Vigencia actualizada correctamente.', 'success')
        db.close()
        return redirect(url_for('vigencias_list'))
    db.close()
    return render_template('vigencias/form.html', vigencia=v)


@app.route('/admin/vigencias/<int:id>/activar', methods=['POST'])
@solo_admin
def vigencia_activar(id):
    db = get_db()
    v = db.execute('SELECT nombre FROM vigencia WHERE id=?', (id,)).fetchone()
    if v:
        db.execute('UPDATE vigencia SET activa=0')
        db.execute('UPDATE vigencia SET activa=1 WHERE id=?', (id,))
        db.commit()
        session['vigencia_id'] = id
        registrar_auditoria('ACTIVAR', 'Vigencias', f'Vigencia {v["nombre"]} activada')
        flash(f'Vigencia "{v["nombre"]}" activada. Todos los usuarios verán esta vigencia al iniciar sesión.', 'success')
    db.close()
    return redirect(url_for('vigencias_list'))


@app.route('/admin/vigencias/<int:id>/copiar', methods=['POST'])
@solo_admin
def vigencia_copiar(id):
    """Crea una nueva vigencia copiando la estructura de procesos (con ponderación en cero)."""
    try:
        año_nuevo = int(request.form.get('año_nuevo', 0))
    except ValueError:
        flash('Año de destino inválido.', 'danger')
        return redirect(url_for('vigencias_list'))
    if not año_nuevo:
        flash('Debe especificar el año de destino.', 'danger')
        return redirect(url_for('vigencias_list'))
    db = get_db()
    if db.execute('SELECT id FROM vigencia WHERE año=?', (año_nuevo,)).fetchone():
        flash(f'Ya existe una vigencia para el año {año_nuevo}.', 'danger')
        db.close()
        return redirect(url_for('vigencias_list'))
    v_origen = db.execute('SELECT * FROM vigencia WHERE id=?', (id,)).fetchone()
    if not v_origen:
        flash('Vigencia origen no encontrada.', 'danger')
        db.close()
        return redirect(url_for('vigencias_list'))
    cur = db.execute('''
        INSERT INTO vigencia (año, nombre, descripcion, activa, fecha_inicio)
        VALUES (?,?,?,0,?)
    ''', (año_nuevo, f'GICA {año_nuevo}',
          f'Copiada de {v_origen["nombre"]}', f'{año_nuevo}-01-01'))
    new_vid = cur.lastrowid
    db.execute('''
        INSERT INTO ponderacion_proceso (proceso_id, vigencia_id, proc_score, ind_score, car_score, norm_score, risk_score)
        SELECT proceso_id, ?, 0, 0, 0, 0, 0
        FROM ponderacion_proceso WHERE vigencia_id=?
        ON CONFLICT(proceso_id, vigencia_id) DO NOTHING
    ''', (new_vid, id))
    db.commit()
    registrar_auditoria('COPIAR', 'Vigencias',
                        f'Vigencia {v_origen["nombre"]} copiada a {año_nuevo}')
    flash(f'Vigencia GICA {año_nuevo} creada con la estructura de {v_origen["nombre"]}.', 'success')
    db.close()
    return redirect(url_for('vigencias_list'))


# ─── COMPARATIVO ─────────────────────────────────────────────────────────────

@app.route('/comparativo')
@login_required
def comparativo():
    db = get_db()
    vigencias = db.execute('SELECT * FROM vigencia ORDER BY año DESC').fetchall()
    vid1 = request.args.get('v1', type=int)
    vid2 = request.args.get('v2', type=int)
    datos = []
    if vid1 and vid2 and vid1 != vid2:
        procesos_list = db.execute(
            'SELECT id, codigo, nombre FROM proceso ORDER BY codigo'
        ).fetchall()
        for p in procesos_list:
            p1 = db.execute(
                'SELECT * FROM ponderacion_proceso WHERE proceso_id=? AND vigencia_id=?',
                (p['id'], vid1)
            ).fetchone()
            p2 = db.execute(
                'SELECT * FROM ponderacion_proceso WHERE proceso_id=? AND vigencia_id=?',
                (p['id'], vid2)
            ).fetchone()
            if p1 or p2:
                pond1 = ((p1['proc_score']+p1['ind_score']+p1['car_score']+
                          p1['norm_score']+p1['risk_score'])/5) if p1 else 0
                pond2 = ((p2['proc_score']+p2['ind_score']+p2['car_score']+
                          p2['norm_score']+p2['risk_score'])/5) if p2 else 0
                datos.append({
                    'proceso': dict(p),
                    'v1': dict(p1) if p1 else None,
                    'v2': dict(p2) if p2 else None,
                    'pond1': round(pond1 * 100, 1),
                    'pond2': round(pond2 * 100, 1),
                    'delta': round((pond2 - pond1) * 100, 1),
                })
    v1_obj = db.execute('SELECT * FROM vigencia WHERE id=?', (vid1,)).fetchone() if vid1 else None
    v2_obj = db.execute('SELECT * FROM vigencia WHERE id=?', (vid2,)).fetchone() if vid2 else None
    db.close()
    return render_template('comparativo.html', vigencias=vigencias, datos=datos,
                           v1=v1_obj, v2=v2_obj, vid1=vid1, vid2=vid2)


@app.route('/api/evolucion-data')
@login_required
def api_evolucion_data():
    db = get_db()
    proceso_id = request.args.get('proceso_id', type=int)
    if proceso_id:
        rows = db.execute('''
            SELECT v.año, v.nombre,
                   COALESCE((pp.proc_score+pp.ind_score+pp.car_score+
                              pp.norm_score+pp.risk_score)/5.0, 0)*100 as ponderacion
            FROM vigencia v
            LEFT JOIN ponderacion_proceso pp ON pp.vigencia_id=v.id AND pp.proceso_id=?
            ORDER BY v.año
        ''', (proceso_id,)).fetchall()
    else:
        rows = db.execute('''
            SELECT v.año, v.nombre,
                   COALESCE(AVG((pp.proc_score+pp.ind_score+pp.car_score+
                                 pp.norm_score+pp.risk_score)/5.0), 0)*100 as ponderacion
            FROM vigencia v
            LEFT JOIN ponderacion_proceso pp ON pp.vigencia_id=v.id
            GROUP BY v.id, v.año, v.nombre
            ORDER BY v.año
        ''').fetchall()
    db.close()
    return jsonify([{
        'año': r['año'],
        'nombre': r['nombre'],
        'ponderacion': round(r['ponderacion'] or 0, 1)
    } for r in rows])


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', '0') == '1'
    app.run(host='0.0.0.0', port=port, debug=debug)
